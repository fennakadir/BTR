import xlrd
import pandas as pd
import glob
from openpyxl import load_workbook, workbook
import os

#script for the extraction of transcripts from the trinotate report based on GO terms of interest
#by Fenna Kadir march 2020

#function to put go terms from excel sheet into a list
#arguments: pathname is path to excel sheet
#           sheetname is the name of the sheet in the excel file
def function1(pathname, sheetname):
    workbook = xlrd.open_workbook(pathname)
    worksheet = workbook.sheet_by_name(sheetname)

    l = []

    for i in range(25):
        if worksheet.cell(i, 0).value == xlrd.empty_cell.value:
            continue
        x = worksheet.cell(i, 0).value
        l.append(x)

    return l

#function to take the go terms from the list and extract the relevant transcripts
#arguments: list is the output from function 1
#           pathfile is the path to the trinotate report
def function2(list, pathfile):
    df = pd.read_excel(pathfile)

    all_data = pd.DataFrame()

    #extract data from trinotate into separate excel files
    for col_num, data in enumerate(list):
        bla = df[df['gene_ontology_blast'].str.contains(data) == True]
        bla.to_excel(r'/Users/fenna/Documents/BTR/code/extraction' + str(col_num) + '.xlsx', index=False)
        excel = load_workbook(r'/Users/fenna/Documents/BTR/code/extraction' + str(col_num) + '.xlsx')
        sheets = excel.sheetnames
        sheet1 = excel[sheets[0]]
        sheet1.cell(row=2, column=17).value = data
        excel.save(r'/Users/fenna/Documents/BTR/code/extraction-' + str(col_num) + '.xlsx')

    #append all different excel files into object
    for f in glob.glob('/Users/fenna/Documents/BTR/code/extraction-*.xlsx'):
        df = pd.read_excel(f)
        all_data = all_data.append(df, ignore_index=True)

    #save object with all the data to excel file
    all_data.to_excel(r'/Users/fenna/Documents/BTR/data/transcript extraction/Transcipts of interest - phosphate iron transport.xlsx', index=False)

    for i in range(col_num + 1):
        os.remove('/Users/fenna/Documents/BTR/code/extraction-' + str(i) + '.xlsx')
        os.remove('/Users/fenna/Documents/BTR/code/extraction' + str(i) + '.xlsx')


list = function1(r'/Users/fenna/Downloads/Phosphate Ion Transport Updated.xlsx', 'Sheet2')

function2(list, r'/Users/fenna/Documents/BTR/data/Trinotate_report.xlsx')
